import openpyxl
import os
from openpyxl.styles import PatternFill


os.chdir("D:\pythontest\pytest\demo")
print(os.getcwd())
wb = openpyxl.load_workbook("guichi.xlsx")
'''wbchizhoushi = openpyxl.load_workbook("池州市.xlsx")
wschizhou = wbchizhoushi.worksheets[0]
'''

ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]
lst1 = []
lst2 = []
orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")


for i in ws2['b']:

    for j in ws1['b']:
        j = j.value
        if j == i.value:
            print(i)
            i.fill = orange_fill


'''ws2.cell(1, 1).fill = orange_fill'''
# wb.save("guichi.xlsx")
