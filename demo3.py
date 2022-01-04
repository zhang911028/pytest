from openpyxl import Workbook
import openpyxl
import os
from openpyxl.drawing.image import Image


os.chdir("D:\pythontest\pytest\demo")
# wb = Workbook()
wb = openpyxl.load_workbook("Mysheet.xlsx")
ws = wb.active

# ws1 = wb.create_sheet("Mysheet")
# print(wb.sheetnames)
'''for i in wb:
    print(i.title)'''

#访问多个单元格
'''print(ws['a1':'c2'])
print(ws['b'])
print(ws[1:10])'''

'''for row in ws.iter_rows(min_row=1, max_row=3, max_col=4):
    for cell in row:
        print(cell)

for col in ws.iter_cols(min_row=1, max_row=3, max_col=2):
    for cell in col:
        print(cell)'''

# 读行或者列
'''for cell in ws.rows:
    print(cell)

for cell in ws.columns:
    print(cell)'''

# 访问单元格仅值
'''for row in ws.values:
    for value in row:
        print(value)'''

'''for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3, values_only=True):
    print(row)'''

# 数据存储
'''for row in ws.iter_rows(min_row=1, max_row=5, max_col=3):
    for cell in row:
        cell.value = 'hello, world'
        print(cell.value)'''

# 插入图像
'''img = Image('pic.png')
ws.add_image(img, 'A1')'''


wb.save("Mysheet.xlsx")
wb.close()
