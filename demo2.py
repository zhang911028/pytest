import openpyxl
from openpyxl.styles import PatternFill
import os


os.chdir("D:\pythontest\pytest\demo")
print(os.getcwd())

wb = openpyxl.load_workbook("guichi.xlsx")
ws = wb.worksheets[0]

lst = []

orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")

with open("demoshiji.csv", mode='r') as f:
    ls = f.readlines()
    for i in ls:
        i = i.replace('\'', '')
        lst.append(i)
'''with open("demoshiji.csv", mode='a', encoding="gbk") as f:
    for row in ws.rows:
        s = [i.value[1] for i in row][2]
        lst.append(s)

        f.write(f'\'{s}\n')'''

for n in ws['b']:
    data = n.value

    for i in lst:
        if i != data:
            n.fill = orange_fill



wb.save("guichi.xlsx")
