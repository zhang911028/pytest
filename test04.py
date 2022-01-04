from openpyxl import Workbook
import openpyxl
import os
from openpyxl.styles import PatternFill


os.chdir(r"D:\pythontest\pytest\demo")
orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
# 创建表
wbwr = Workbook(write_only=True)
wswr = wbwr.create_sheet("writedata")

print(wswr)


# 读取本地表
wbrd = openpyxl.load_workbook("Mysheet.xlsx", read_only=True)
ws1 = wbrd["Sheet"]
ws2 = wbrd['Mysheet']

lst = []
# 分析表
for row1 in ws2.values:
    # print(row[0])
    card1 = row1[0]

    for row in ws1.values:
        # print(row[1])
        card2 = row[1]
        if card1 == card2:
            wswr.append(row)

wbrd.close()


wbwr.save("writedata.xlsx")
wbwr.close()
